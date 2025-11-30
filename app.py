
import os
import json
import logging
import base64
from io import BytesIO
from typing import Tuple

import requests
from flask import Flask, request, jsonify
import msal
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
import re

# ----------------------------
# Config & Logging
# ----------------------------
logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
logger = logging.getLogger(__name__)

TENANT_ID = os.getenv("TENANT_ID")
CLIENT_ID = os.getenv("CLIENT_ID")
CLIENT_SECRET = os.getenv("CLIENT_SECRET")
AUTHORITY = f"https://login.microsoftonline.com/{TENANT_ID}" if TENANT_ID else None
SCOPES = ["https://graph.microsoft.com/.default"]
GRAPH_BASE = "https://graph.microsoft.com/v1.0"

app = Flask(__name__)

# ----------------------------
# MSAL / Token helpers
# ----------------------------
_msal_app = None

def _get_msal_app():
    global _msal_app
    if _msal_app is None:
        if not (TENANT_ID and CLIENT_ID and CLIENT_SECRET):
            raise RuntimeError("Faltan variables de entorno TENANT_ID/CLIENT_ID/CLIENT_SECRET")
        _msal_app = msal.ConfidentialClientApplication(
            CLIENT_ID,
            authority=AUTHORITY,
            client_credential=CLIENT_SECRET,
        )
    return _msal_app


def get_graph_token() -> str:
    app_msal = _get_msal_app()
    result = app_msal.acquire_token_silent(SCOPES, account=None)
    if not result:
        result = app_msal.acquire_token_for_client(scopes=SCOPES)
    if "access_token" not in result:
        logger.error(f"Error obteniendo token Graph: {result}")
        raise RuntimeError("No se pudo obtener token de Graph")
    return result["access_token"]

# ----------------------------
# Graph helpers (shares, download, upload)
# ----------------------------
def encode_sharing_url(url: str) -> str:
    """Graph /shares/{encoded}/driveItem requiere URL en base64 url-safe con prefijo 'u!'."""
    b64 = base64.b64encode(url.encode("utf-8")).decode("utf-8")
    b64 = b64.replace("/", "_").replace("+", "-").rstrip("=")
    return f"u!{b64}"

def resolve_share_to_item(url: str, token: str) -> Tuple[str, str]:
    """Devuelve (drive_id, item_id) a partir de un sharing/webUrl de OneDrive/SharePoint."""
    encoded = encode_sharing_url(url)
    headers = {"Authorization": f"Bearer {token}"}
    r = requests.get(f"{GRAPH_BASE}/shares/{encoded}/driveItem", headers=headers)
    if r.status_code != 200:
        logger.error(f"Resolve share fallo ({r.status_code}): {r.text}")
        raise RuntimeError("No se pudo resolver el enlace compartido (shares/driveItem)")
    data = r.json()
    drive_id = data.get("parentReference", {}).get("driveId")
    item_id = data.get("id")
    if not drive_id or not item_id:
        raise RuntimeError("Faltan driveId/itemId en la respuesta de driveItem")
    return drive_id, item_id

def download_item_content(drive_id: str, item_id: str, token: str) -> bytes:
    headers = {"Authorization": f"Bearer {token}"}
    url = f"{GRAPH_BASE}/drives/{drive_id}/items/{item_id}/content"
    r = requests.get(url, headers=headers)
    if r.status_code not in (200, 302):
        logger.error(f"Descarga fallo ({r.status_code}): {r.text}")
        raise RuntimeError("No se pudo descargar el contenido del archivo")
    return r.content

def upload_item_content(drive_id: str, item_id: str, content: bytes, token: str) -> None:
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/octet-stream"}
    url = f"{GRAPH_BASE}/drives/{drive_id}/items/{item_id}/content"
    r = requests.put(url, headers=headers, data=content)
    if r.status_code not in (200, 201):
        logger.error(f"Upload fallo ({r.status_code}): {r.text}")
        raise RuntimeError("No se pudo subir (sobrescribir) el archivo destino")

# ----------------------------
# Excel helpers: anchor search, formula shifting, copy range
# ----------------------------
def normalize_text(s: str) -> str:
    return (s or "").strip().replace(":", "").casefold()

def find_anchor_row(ws, col_letter: str, search_text: str) -> int:
    target = normalize_text(search_text)
    col_idx = column_index_from_string(col_letter)
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(row=r, column=col_idx)
        if normalize_text(cell.value) == target:
            return r
    # Fallback: C por si cambió el merge
    if col_letter.upper() != "C":
        col_idx = column_index_from_string("C")
        for r in range(1, ws.max_row + 1):
            cell = ws.cell(row=r, column=col_idx)
            if normalize_text(cell.value) == target:
                return r
    raise ValueError(f"No se encontró '{search_text}' en columna {col_letter} ni en C.")

# Regex que soporta prefijo de hoja opcional y rangos
_token = r"(?:'[^']+'|[A-Za-z0-9_]+)!"  # 'Hoja 1'!  o Hoja1!
sheet_prefix = rf"(?:{_token})?"
cell_ref = r"(\$?)([A-Z]{1,3})(\$(\d+)|(\d+))"  # col_abs, col_letters, row_part, num1?, num2?
cell_or_range = rf"{sheet_prefix}{cell_ref}(?:: {sheet_prefix}{cell_ref})?"
cell_or_range_pattern = re.compile(cell_or_range.replace(" ", ""))

def _shift_one(groups, drow, dcol):
    col_abs, col_letters, row_part, row_num1, row_num2 = groups
    row_abs = "$" if row_part.startswith("$") else ""
    row_number = int(row_num1 or row_num2)
    col_idx = column_index_from_string(col_letters)
    new_col_idx = col_idx if col_abs == "$" else col_idx + dcol
    new_row = row_number if row_abs == "$" else row_number + drow
    new_col_idx = max(1, new_col_idx)
    new_row = max(1, new_row)
    return f"{col_abs}{get_column_letter(new_col_idx)}{row_abs}{new_row}"

def shift_formula_refs(formula: str, drow: int = 0, dcol: int = 0) -> str:
    """Desplaza referencias en una fórmula Excel. Respeta textos "..." y no toca INDIRECT/ADDRESS."""
    if not (isinstance(formula, str) and formula.startswith("=")):
        return formula
    up = formula.upper()
    if "INDIRECT" in up or "ADDRESS" in up:
        return formula

    def repl(m):
        text = m.group(0)
        # Evitar tocar coincidencias dentro de comillas
        before = formula[:m.span()[0]]
        if before.count('"') % 2 == 1:
            return text
        g = m.groups()
        if len(g) == 10:  # rango: dos refs
            left = _shift_one(g[1:6], drow, dcol)
            right = _shift_one(g[6:], drow, dcol)
            return f"{left}:{right}"
        else:  # ref simple
            return _shift_one(g, drow, dcol)

    return cell_or_range_pattern.sub(repl, formula)

def copy_range_adjusting(ws_src, ws_dst, src_range: str, dst_start_cell: str):
    start, end = src_range.split(":")
    src_start_col_letter = ''.join(filter(str.isalpha, start))
    src_start_row = int(''.join(filter(str.isdigit, start)))
    src_end_col_letter = ''.join(filter(str.isalpha, end))
    src_end_row = int(''.join(filter(str.isdigit, end)))
    src_start_col = column_index_from_string(src_start_col_letter)
    src_end_col = column_index_from_string(src_end_col_letter)
    num_rows = src_end_row - src_start_row + 1
    num_cols = src_end_col - src_start_col + 1

    dst_start_col_letter = ''.join(filter(str.isalpha, dst_start_cell))
    dst_start_row = int(''.join(filter(str.isdigit, dst_start_cell)))
    dst_start_col = column_index_from_string(dst_start_col_letter)

    drow = dst_start_row - src_start_row
    dcol = dst_start_col - src_start_col

    for r_off in range(num_rows):
        for c_off in range(num_cols):
            r_src = src_start_row + r_off
            c_src = src_start_col + c_off
            r_dst = dst_start_row + r_off
            c_dst = dst_start_col + c_off

            src = ws_src.cell(row=r_src, column=c_src)
            dst = ws_dst.cell(row=r_dst, column=c_dst)

            val = src.value
            if isinstance(val, str) and val.startswith("="):
                dst.value = shift_formula_refs(val, drow=drow, dcol=dcol)
            else:
                dst.value = val

            # Copiar estilos básicos
            dst.font = src.font
            dst.fill = src.fill
            dst.border = src.border
            dst.alignment = src.alignment
            dst.number_format = src.number_format

    # Copiar anchos de columna
    for j in range(num_cols):
        src_letter = get_column_letter(src_start_col + j)
        dst_letter = get_column_letter(dst_start_col + j)
        dim = ws_src.column_dimensions.get(src_letter)
        if dim and dim.width:
            ws_dst.column_dimensions[dst_letter].width = dim.width

# ----------------------------
# Flask endpoints
# ----------------------------
@app.get("/health")
def health():
    return jsonify({"status": "ok"})

@app.post("/copy-range")
def copy_range_endpoint():
    try:
        body = request.get_json(force=True)
        logger.info(f"Body recibido: {json.dumps(body, ensure_ascii=False)}")

        source_sharing_url = body["source_sharing_url"]
        source_sheet = body["source_sheet"]
        source_range = body["source_range"]
        dest_sharing_url = body["dest_sharing_url"]
        dest_sheet = body["dest_sheet"]
        search_col_letter = body.get("search_col_letter", "B")
        search_text = body["search_text"]
        offset_rows = int(body.get("offset_rows", 0))

        token = get_graph_token()

        # Resolver y descargar origen
        src_drive, src_item = resolve_share_to_item(source_sharing_url, token)
        src_bytes = download_item_content(src_drive, src_item, token)
        wb_src = load_workbook(BytesIO(src_bytes), data_only=False)
        if source_sheet not in wb_src.sheetnames:
            raise RuntimeError(f"La hoja origen '{source_sheet}' no existe.")
        ws_src = wb_src[source_sheet]

        # Resolver, descargar destino
        dst_drive, dst_item = resolve_share_to_item(dest_sharing_url, token)
        dst_bytes = download_item_content(dst_drive, dst_item, token)
        wb_dst = load_workbook(BytesIO(dst_bytes), data_only=False)
        if dest_sheet not in wb_dst.sheetnames:
            raise RuntimeError(f"La hoja destino '{dest_sheet}' no existe.")
        ws_dst = wb_dst[dest_sheet]

        # Buscar ancla y calcular punto de pegado
        anchor_row = find_anchor_row(ws_dst, col_letter=search_col_letter, search_text=search_text)
        paste_start_row = anchor_row + offset_rows
        dst_start_cell = f"B{paste_start_row}"  # pegamos en columna B

        # Copiar rango ajustando fórmulas/formatos/anchos
        copy_range_adjusting(ws_src, ws_dst, src_range=source_range, dst_start_cell=dst_start_cell)

        # Guardar y subir destino
        out = BytesIO()
        wb_dst.save(out)
        upload_item_content(dst_drive, dst_item, out.getvalue(), token)

        # Calcular filas/columnas del rango para la respuesta
        start, end = source_range.split(":")
        rows = int(''.join(filter(str.isdigit, end))) - int(''.join(filter(str.isdigit, start))) + 1
        cols = column_index_from_string(''.join(filter(str.isalpha, end))) - column_index_from_string(''.join(filter(str.isalpha, start))) + 1

        return jsonify({
            "status": "ok",
            "paste_start": dst_start_cell,
            "rows": rows,
            "cols": cols
        })

    except ValueError as ve:
        logger.error(f"BadRequest: {ve}")
        return jsonify({"error": str(ve)}), 400
    except Exception as e:
        logger.exception("Error en /copy-range")
        return jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    port = int(os.getenv("PORT", "10000"))
    app.run(host="0.0.0.0", port=port)
